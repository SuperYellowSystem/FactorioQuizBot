# ======================================================================
# imports
# ======================================================================
from openpyxl import load_workbook

import datetime
from tzlocal import get_localzone

from discord import Embed

import random
import os

import logging
logger = logging.getLogger(__name__)

# ======================================================================
solution_index = {
    1: '\U0001F1E6',
    2: '\U0001F1E7',
    3: '\U0001F1E8',
    4: '\U0001F1E9'
}


class QuizItem:
    """
    This class represent an item(question) of the quiz.
    """
    def __init__(self, i18n, qid, question, answers, solution, difficulty, domain):
        """
        Create a simple item with question/answer information

        :param qid: the question id
        :param question: the question itself
        :param answers: an array with multiple answers
        :param solution: the index of the correct answer
        :param difficulty: the difficulty of the question
        :param domain: the theme the question belong
        """
        self.i18n = i18n
        self._qid = qid
        self._question = question
        self._answers = answers
        self._solution = solution
        self._difficulty = difficulty
        self._domain = domain

    def create_item_embed(self):
        time = datetime.datetime.now(get_localzone())
        embed = Embed(colour=3447003, timestamp=time, title=f'{self._qid}/ {self._domain}')
        answer_code = ord('A')
        output_answers = ""
        for answer in self._answers:
            output_answers += f'{chr(answer_code)}: {answer}     '
            answer_code += 1
        embed.add_field(name=f'{self._question}', value=output_answers)
        embed.set_footer(text=f'{self.i18n.cmdQuizList_difficulty}: {self._difficulty}')
        return embed

    def get_solution(self):
        return solution_index.get(self._solution)

    def get_nb_answers(self):
        return len(self._answers)

    def get_difficulty(self):
        return self._difficulty


# ======================================================================
class QuizList:
    """
    This class generate and store all the items(questions) for the quiz
    """
    def __init__(self, i18n, is_exam):
        """
        Generate a list of item from a xlsx file and store it in a list
        """
        self.__item_list = []
        self.__index = -1

        # index for difficulty
        self.__diff1 = None
        self.__diff2 = None

        path = os.path.dirname(os.path.abspath(__file__))
        if os.name == 'nt':
            index = path.find("\\cogs")
        else:
            index = path.find("/cogs")
        file = f'{path[:index]}/questions/QCM_Factorio_{i18n.get_language()}.xlsx'
        wb = load_workbook(filename=file, read_only=True)
        ws = wb['Quiz']
        for row in range(1, ws.max_row):
            qid = ws.cell(row=row + 1, column=1).value
            question = ws.cell(row=row + 1, column=2).value
            answer1 = ws.cell(row=row + 1, column=3).value
            answer2 = ws.cell(row=row + 1, column=4).value
            answer3 = ws.cell(row=row + 1, column=5).value
            answer4 = ws.cell(row=row + 1, column=6).value
            solution = ws.cell(row=row + 1, column=7).value
            answer_total = ws.cell(row=row + 1, column=8).value
            difficulty = ws.cell(row=row + 1, column=9).value
            domain = ws.cell(row=row + 1, column=10).value
            if self.__diff1 is None and difficulty == 2:
                self.__diff1 = qid - 1
            elif self.__diff2 is None and difficulty == 3:
                self.__diff2 = qid - 1

            answers = [answer1, answer2]
            if answer_total > 2:
                answers.append(answer3)
            if answer_total > 3:
                answers.append(answer4)
            self.__item_list.append(QuizItem(i18n, qid, question, answers, solution, difficulty, domain))

        if is_exam:
            self.__item_list = [*random.sample(self.__item_list[:self.__diff1], 6),
                                *random.sample(self.__item_list[self.__diff1:self.__diff2], 7),
                                *random.sample(self.__item_list[self.__diff2:], 7)]

    def select_next_item(self):
        self.__index += 1
        if self.__index == len(self.__item_list):
            self.__index = 0
        return self.__item_list[self.__index]

    def shuffle_list(self):
        random.shuffle(self.__item_list)
        self.__index = -1

# ======================================================================
# Test
# ======================================================================
# if __name__ == '__main__':
